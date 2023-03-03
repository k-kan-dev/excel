
import logging
import os
import re
from typing import Any, Dict, List, Set, Tuple, Union

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.cell import Cell



from table import Table
from abc.excel import ABC_Excel


class ExcelBase(ABC_Excel):
    def __init__(self,
                 fn: str,
                 sheet: Union[int, str] = 0,
                 ):
        self.fn = str(fn)
        if fn:
            self.f = load_workbook(
                filename=self.fn,
                data_only=True
            )

            ### get sheet obj
            if type(sheet) == str:
                for sheet_name in self.f.sheetnames:
                    if re.search(rf"{sheet}", sheet_name):
                        self.sheet_name = sheet_name
                        self.target_error_code = sheet_name.replace("テストパターン_", "")
                        self.sheet = self.f[str(sheet_name)]
                        break
            else:
                self.sheet = self.f.worksheets[sheet]

            ### get all merge cell
            self.mergedcellranges = self.sheet.merged_cells.ranges

    def get_head_cell_in_merged_cell(self,
                                     merged_cell: MergedCell):
        """_summary_

        Args:
            merged_cell (MergedCell): _description_

        Returns:
            _type_: _description_
        """
        for mcr in self.mergedcellranges:
            if (merged_cell.row, merged_cell.column) in list(mcr.cells):
                (head_row, head_column) = list(mcr.cells)[0]
                head_cell = self.sheet.cell(row=head_row, column=head_column)
                return head_cell

    def get_neighbor_cell(self, cell, x, y):
        """_summary_
        """
        return self.sheet.cell(row=cell.row+x, column=cell.column+y)

    def chg_font_color(self,
                       rs: int,
                       cs: int,
                       re: int, 
                       ce: int,
                       *,
                       all: bool = False,
                       column: List[int] = None,
                       row: List[int] = None,
                       font_name: str = "ＭＳ ゴシック",
                       font_size: int = 10,
                       font_color: str = "FF0000",
                       ):
        """ assin range(rs,cs,re,ce), 

        Args:
            rs (int): the row number as start point
            cs (int): the column number as start point
            re (int): the row number as end point
            ce (int): the column number as end point
            column (List[int]): columns to change font. Defaults to None.
            row (List[int]): rows to change font. Defaults to None.
            font_name (str): name. Defaults to 'ＭＳ ゴシック'.
            font_size (int): size. Defaults to 10.
            font_color (str): color. Defaults to 'FF0000'.
            all (bool): command to change font of all cells. Defaults to False.
        """
        if column:
            for c in column:
                for cell in list(self.sheet.columns)[c][rs-1:re]:
                    cell.font = Font(name=font_name,
                                     color=font_color,
                                     size=font_size)
        elif row:
            for r in row:
                for cell in list(self.sheet.rows)[r][cs-1:ce]:
                    cell.font = Font(name=font_name,
                                     color=font_color,
                                     size=font_size)
        elif all:
            for c in range(ce)[cs-1:]:
                for cell in list(self.sheet.columns)[c][rs-1:re]:
                    cell.font = Font(name=font_name,
                                     color=font_color,
                                     size=font_size)
        else:
            raise("column, row, or all should be chosen.")

    def save(self,
             file_affix: str,
             output_path: str):
        save_fname = self.fn.replace("resource\\", "")
        save_fname = save_fname.replace("output\\after\\", "")
        save_fname = save_fname.replace("_after", "")

        if file_affix:
            f = save_fname.split(".")
            save_fname = f[0] + "_" + file_affix +"." + f[1]

        save_fname = os.path.join(
            output_path, 
            save_fname
        )

        os.makedirs(os.path.dirname(save_fname),
                    exist_ok=True)
        self.f.save(save_fname)
        print(f"saved as : {save_fname}")
